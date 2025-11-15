import openpyxl

def read_lifting_parameters(filename="lifting_plan.xlsx"):
    try:
        workbook = openpyxl.load_workbook(filename, data_only=True)
        sheet = workbook["LiftingPlan"]
        
        params = {}
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            param_name = row[0].value
            param_value = row[1].value
            if param_name:
                params[param_name.strip()] = param_value
        
        return params

    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return None

if __name__ == '__main__':
    parameters = read_lifting_parameters()
    if parameters:
        for key, value in parameters.items():
            print(f"{key}: {value}")
