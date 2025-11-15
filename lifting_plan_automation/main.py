
from lifting_plan_automation import excel_io, calculations, drawing
import openpyxl
import os

def main(filepath="lifting_plan.xlsx", output_dir=".", load_weight=None, sling_angle=None, num_slings=None, crane_capacity=None, sling_swl=None):
    params = excel_io.read_lifting_parameters(filepath)
    if params is None:
        params = {}

    # Override params with values from arguments if they are provided
    if load_weight is not None:
        params["Load Weight"] = load_weight
    if num_slings is not None:
        params["Number of Slings"] = num_slings
    if sling_swl is not None:
        params["Sling SWL"] = sling_swl
    if crane_capacity is not None:
        params["Crane Capacity"] = crane_capacity

    load_weight = params.get("Load Weight", 0)
    load_length = params.get("Load Length", 0)
    load_width = params.get("Load Width", 0)
    num_slings = params.get("Number of Slings", 0)
    sling_length = params.get("Sling Length", 0)
    sling_swl = params.get("Sling SWL", 0)
    shackle_swl = params.get("Shackle SWL", 0)
    hook_weight = params.get("Hook Weight", 0)
    crane_capacity = params.get("Crane Capacity", 0)

    cog_x, cog_y = calculations.calculate_center_of_gravity(load_length, load_width)
    if sling_angle is None:
        sling_angle = calculations.calculate_sling_angle(sling_length, load_width, num_slings)
    
    sling_tension = calculations.calculate_sling_tension(load_weight, num_slings, sling_angle)
    sling_sf, shackle_sf = calculations.check_safety_factors(sling_tension, sling_swl, shackle_swl)
    crane_utilization = calculations.calculate_crane_utilization(load_weight, hook_weight, crane_capacity)

    results = {
        "Center of Gravity (X)": cog_x,
        "Center of Gravity (Y)": cog_y,
        "Sling Angle": sling_angle,
        "Sling Tension": sling_tension,
        "Sling Safety Factor": sling_sf,
        "Shackle Safety Factor": shackle_sf,
        "Crane Utilization": crane_utilization,
    }

    try:
        # It's possible the file doesn't exist if we're generating from inputs only
        if os.path.exists(filepath):
            workbook = openpyxl.load_workbook(filepath)
        else:
            workbook = openpyxl.Workbook()

        sheet_name = "LiftingPlan"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            sheet.title = sheet_name


        # Create a new dictionary for results to be written to excel, including original params
        excel_output_data = params.copy()
        excel_output_data.update(results)

        # Clear the sheet and write the new data
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(["Parameter", "Value"])
        for key, value in excel_output_data.items():
            sheet.append([key, value])
        
        output_filename_base, _ = os.path.splitext(os.path.basename(filepath))
        output_excel_path = os.path.join(output_dir, f"{output_filename_base}_results.xlsx")
        workbook.save(output_excel_path)
        print(f"Results saved to '{output_excel_path}'")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

    output_dxf_path = os.path.join(output_dir, f"{output_filename_base}.dxf")
    # Update params with results for drawing
    params.update(results)
    drawing.create_lifting_plan_drawing(output_dxf_path, params, results)

    return output_excel_path, output_dxf_path

if __name__ == "__main__":
    main()

